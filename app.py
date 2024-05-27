import openpyxl
from urllib.parse import quote
import webbrowser #abrir navegador
from time import sleep
import pyautogui




webbrowser.open('https://web.whatsapp.com/')
sleep(30)

# Carrega o arquivo Excel
workbook = openpyxl.load_workbook('mensagens_teste.xlsx')

# Lista todas as planilhas no arquivo
print("Nomes das planilhas:", workbook.sheetnames)



# Acessa a planilha com o nome correto
pagina_cliente = workbook['Planilha1']

for cliente in pagina_cliente.iter_rows(min_row=2):
    nome = cliente[0].value
    telefone = cliente[1].value
    mensagem_personalizada =  f'Olá, {nome} meu querido (a). Muito boa noite. Isso é um teste de automação!!!. O Sport perdeu foi? '
    
    try:
        link_mens_zap = f'https://web.whatsapp.com/send?phone={telefone}&text={quote(mensagem_personalizada)}'
        webbrowser.open(link_mens_zap)
        sleep(10)
        seta = pyautogui.locateCenterOnScreen('seta.png')
        sleep(5)
        pyautogui.click(seta[0], seta[1])
        sleep(5)
        pyautogui.hotkey('crtl','w')
        sleep(5)
    except Exception as e:
        print(f'Não foi possivel passar mensagem para {nome}')
        with open('erros.csv', 'a', newline='', encoding='utf-8') as arquivo:
            arquivo.write(f'{nome}, {telefone}, {e}')

print('fim de progrma')